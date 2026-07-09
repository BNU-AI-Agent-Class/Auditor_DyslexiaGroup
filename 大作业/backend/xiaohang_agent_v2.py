from dotenv import load_dotenv; load_dotenv()
from openai import OpenAI
import os
import sys
import json
import re
import random
from datetime import datetime
from collections import defaultdict

# ------------------------------------------------------------------
# 路径与配置
# ------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
memory_file = os.path.join(BASE_DIR, "agent.md")
skill_crisis_file = os.path.join(BASE_DIR, "skill_危机转介.md")
history_file = os.path.join(BASE_DIR, "xiaohang_history_v2.json")

# 数据文件路径（优先从环境变量，否则用相对路径）
_DATA_ROOT = os.getenv("DATA_DIR", os.path.join(BASE_DIR, "..", "data"))
DATA_DIR = os.path.join(_DATA_ROOT, "覆盖率&识读率网页版")
CHAR_TABLE_FILE = os.path.join(DATA_DIR, "3600字2020册数.xlsx")
FREQ_TABLE_FILE = os.path.join(DATA_DIR, "平衡语料库字频表计算用.xlsx")
PROP_TABLE_FILE = os.path.join(DATA_DIR, "抽取比例设置.xlsx")
TEXT_LIBRARY_DIR = os.path.join(_DATA_ROOT, "识读率文本")

api_key = os.getenv("DEEPSEEK_API_KEY")
if not api_key:
    print("[错误] 找不到 DEEPSEEK_API_KEY。请在 .env 文件中设置：")
    print("       DEEPSEEK_API_KEY=sk-xxxxxxxxxxxx")
    sys.exit(1)

model = os.getenv("DEEPSEEK_MODEL", "deepseek-chat")
client = OpenAI(api_key=api_key, base_url="https://api.deepseek.com")

# 学期名称
SEMESTER_NAMES = ['一上', '一下', '二上', '二下', '三上', '三下',
                  '四上', '四下', '五上', '五下', '六上', '六下']
UNCATEGORIZED = 99

# 推荐阈值（来自 PPT）
PASS_THRESHOLD = 0.95

# 终止条件
PAUSE_AFTER_CONSECUTIVE_ERRORS = 10
STOP_AFTER_TOTAL_ERRORS = 15

# 等级映射：学期 -> 等级编码（对应文件夹前缀）
SEMESTER_TO_LEVEL = {
    1: '3',   # 一上 -> 等级3（偏难提示）
    2: '3',   # 一下 -> 等级3（偏难提示）
    3: '3',   # 二上 -> 等级3
    4: '4',   # 二下 -> 等级4
    5: '5',   # 三上 -> 等级5
    6: '6',   # 三下 -> 等级6（默认目标用户）
    7: '6',   # 四上 -> 等级6（偏简单提示）
    8: '6',   # 四下 -> 等级6（偏简单提示）
    9: '6',   # 五上 -> 等级6（偏简单提示）
    10: '6',  # 五下 -> 等级6（偏简单提示）
    11: '6',  # 六上 -> 等级6（偏简单提示）
    12: '6',  # 六下 -> 等级6（偏简单提示）
}

# 主题映射：口语化主题 -> {等级: 文件夹名}
# 文件夹命名格式：{等级}{主题}
THEME_OPTIONS = {
    "童话故事和诗歌": {
        '5': '5-1童话&诗歌',
    },
    "生活中的小故事和神话寓言": {
        '3': '3生活故事&神话寓言',
        '4': '4生活故事&神话寓言',
        '5': '5-2生活故事&神话寓言',
    },
    "生活故事和散文": {
        '5': '5-3生活故事&散文',
        '6': '6-1生活故事&散文',
    },
    "生活故事和历史故事": {
        '6': '6-2生活故事&历史故事',
    },
}

# 反向查找：文件夹名 -> 口语化主题
FOLDER_TO_THEME = {}
for theme, levels in THEME_OPTIONS.items():
    for level, folder in levels.items():
        FOLDER_TO_THEME[folder] = theme

# ------------------------------------------------------------------
# 工具函数
# ------------------------------------------------------------------
def load_prompt():
    """加载 agent.md 作为系统提示词。"""
    if not os.path.exists(memory_file):
        print(f"[错误] 找不到 {memory_file}，请先创建人设文件。")
        sys.exit(1)
    with open(memory_file, "r", encoding="utf-8") as f:
        return f.read()


def load_skill_crisis():
    """加载 skill_危机转介.md，解析红灯/黄灯关键词。"""
    if not os.path.exists(skill_crisis_file):
        print(f"[警告] 找不到 {skill_crisis_file}，使用默认危机检测。")
        return None

    with open(skill_crisis_file, "r", encoding="utf-8") as f:
        content = f.read()

    # 解析红灯关键词（简化：提取引号中的内容）
    crisis_keywords = []
    yellow_keywords = []

    lines = content.split('\n')
    current_section = None

    for line in lines:
        if '🔴 红灯信号' in line:
            current_section = 'red'
        elif '🟡 黄灯信号' in line:
            current_section = 'yellow'
        elif '示例关键词' in line or '|' in line:
            continue  # 跳过表头行
        elif line.strip().startswith('"') and current_section:
            # 提取引号中的关键词
            import re
            keywords = re.findall(r'"([^"]+)"', line)
            if current_section == 'red':
                crisis_keywords.extend(keywords)
            elif current_section == 'yellow':
                yellow_keywords.extend(keywords)
        elif '自我否定' in line or '生命危机' in line:
            current_section = 'red'
        elif '挫败' in line or '放弃' in line:
            current_section = 'yellow'

    # 如果解析失败，使用默认列表
    if not crisis_keywords:
        crisis_keywords = [
            "不想活了", "活着没意思", "想死", "自杀",
            "好笨", "我太笨了", "我很笨", "我读不好"
        ]
    if not yellow_keywords:
        yellow_keywords = [
            "不想读了", "不读了", "太难了", "好累", "有点烦", "读得慢"
        ]

    return {
        'crisis_keywords': crisis_keywords,
        'yellow_keywords': yellow_keywords,
        'content': content,
    }


def save_history(messages):
    """保存对话历史。"""
    os.makedirs(os.path.dirname(history_file) or ".", exist_ok=True)
    with open(history_file, "w", encoding="utf-8") as f:
        json.dump(messages, f, ensure_ascii=False, indent=2)


def load_history():
    """加载对话历史。"""
    if os.path.exists(history_file):
        try:
            with open(history_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []


def chat(messages):
    """调用 DeepSeek API。"""
    response = client.chat.completions.create(model=model, messages=messages)
    return response.choices[0].message.content


# ------------------------------------------------------------------
# Excel 数据加载
# ------------------------------------------------------------------
def load_char_table():
    """加载汉字→册数字表。返回 {汉字: 册数索引(0-11) 或 99}。"""
    import openpyxl
    char_map = {}
    wb = openpyxl.load_workbook(CHAR_TABLE_FILE, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        char, semester = row[0], row[1]
        if not char or not isinstance(char, str):
            continue
        char = char.strip()
        if not char:
            continue
        sem = int(semester) if semester is not None else UNCATEGORIZED
        if 1 <= sem <= 12:
            char_map[char] = sem - 1  # 转为 0-11 索引
        else:
            char_map[char] = UNCATEGORIZED
    wb.close()
    return char_map


def load_freq_table():
    """加载汉字→频率表。返回 {汉字: 频率}。"""
    import openpyxl
    freq_map = {}
    wb = openpyxl.load_workbook(FREQ_TABLE_FILE, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        char, freq = row[0], row[1]
        if not char or not isinstance(char, str):
            continue
        char = char.strip()
        if not char:
            continue
        freq_map[char] = float(freq) if freq is not None else 0.0
    wb.close()
    return freq_map


def load_prop_table():
    """加载抽取比例设置表。返回 {(儿童学期, 汉字册数): (比例1, 比例2)}。"""
    import openpyxl
    prop_map = {}
    wb = openpyxl.load_workbook(PROP_TABLE_FILE, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        child_sem, _, target_sem, prop1, prop2 = row
        child_sem = int(child_sem) if child_sem is not None else 0
        target_sem = int(target_sem) if target_sem is not None else 0
        prop1 = float(prop1) if prop1 is not None else 0.0
        prop2 = float(prop2) if prop2 is not None else 0.0
        prop_map[(child_sem, target_sem)] = (prop1, prop2)
    wb.close()
    return prop_map


# ------------------------------------------------------------------
# 文本库加载
# ------------------------------------------------------------------
def load_text_library():
    """加载识读率文本库。返回 {文件夹名: [{name, content}, ...]}。"""
    library = {}
    if not os.path.exists(TEXT_LIBRARY_DIR):
        return library
    for folder_name in sorted(os.listdir(TEXT_LIBRARY_DIR)):
        folder_path = os.path.join(TEXT_LIBRARY_DIR, folder_name)
        if not os.path.isdir(folder_path):
            continue
        texts = []
        for filename in sorted(os.listdir(folder_path)):
            if not filename.endswith('.txt'):
                continue
            file_path = os.path.join(folder_path, filename)
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    content = f.read()
                texts.append({
                    "name": filename[:-4],  # 去掉 .txt
                    "content": content,
                    "folder": folder_name,
                })
            except Exception as e:
                print(f"[警告] 读取文本失败 {file_path}: {e}")
        if texts:
            library[folder_name] = texts
    return library


def extract_readable_title(filename):
    """从文件名中提取给孩子看的标题。"""
    # 文件名格式如：人教-03-21-窗前的气球
    parts = filename.split('-')
    if len(parts) >= 4:
        return '-'.join(parts[3:])  # 取第四段及以后
    if len(parts) == 3:
        return parts[2]
    return filename


def extract_chinese_chars(text):
    """从文本中提取所有汉字。"""
    return [c for c in text if re.match(r'[一-鿿]', c)]


# ------------------------------------------------------------------
# 核心算法：识读率测验组卷
# ------------------------------------------------------------------
def assemble_test(child_semester, texts, char_map, freq_map, prop_map):
    """
    根据儿童学期和10篇文本，生成分层抽样的识读率测验字表。
    返回 {
        'total_chars': int,
        'groups': [{semester, semesterName, sampleSize, totalSource, proportion, chars: [{char, freq}]}],
        'all_chars': [char, ...],
        'child_semester': int,
        'used_scheme': int,
    }
    """
    # Step 1: 提取10篇文本中的所有汉字
    all_text_chars = set()
    for t in texts:
        for c in extract_chinese_chars(t['content']):
            all_text_chars.add(c)

    # Step 2: 按册数分组
    sem_chars = {i: [] for i in range(12)}
    sem_chars[UNCATEGORIZED] = []
    for c in all_text_chars:
        sem = char_map.get(c, UNCATEGORIZED)
        if sem in sem_chars:
            sem_chars[sem].append(c)

    # Step 3: 尝试方案1
    result = _try_scheme(child_semester, sem_chars, prop_map, freq_map, 'prop1')
    used_scheme = 1

    # Step 4: 如果 > 180 字，改用方案2
    if result['totalChars'] > 180:
        result = _try_scheme(child_semester, sem_chars, prop_map, freq_map, 'prop2')
        used_scheme = 2

    return {
        'total_chars': result['totalChars'],
        'groups': result['groups'],
        'all_chars': [item['char'] for g in result['groups'] for item in g['chars']],
        'child_semester': child_semester,
        'used_scheme': used_scheme,
    }


def _try_scheme(child_semester, sem_chars, prop_map, freq_map, prop_key):
    """尝试一种抽样方案。"""
    groups = []
    total = 0

    for i in range(12):
        chars = sem_chars.get(i, [])
        if not chars:
            continue

        rule = prop_map.get((child_semester, i + 1))
        if not rule:
            continue
        proportion = rule[0] if prop_key == 'prop1' else rule[1]
        if proportion <= 0:
            continue

        sample_size = max(1, round(len(chars) * proportion))

        # 按频率排序并分层抽样
        chars_with_freq = [{'char': c, 'freq': freq_map.get(c, 0.0)} for c in chars]
        chars_with_freq.sort(key=lambda x: x['freq'])

        sampled = []
        valid_sample = min(sample_size, len(chars_with_freq))
        for j in range(valid_sample):
            start = int(j * len(chars_with_freq) / valid_sample)
            end = int((j + 1) * len(chars_with_freq) / valid_sample)
            stratum = chars_with_freq[start:end]
            picked = stratum[len(stratum) // 2]
            sampled.append(picked)

        groups.append({
            'semester': i,
            'semesterName': SEMESTER_NAMES[i],
            'sampleSize': len(sampled),
            'totalSource': len(chars),
            'proportion': proportion,
            'chars': sampled,
        })
        total += len(sampled)

    # 处理未收录字（99）
    chars99 = sem_chars.get(UNCATEGORIZED, [])
    if chars99:
        rule = prop_map.get((child_semester, UNCATEGORIZED))
        if rule:
            proportion = rule[0] if prop_key == 'prop1' else rule[1]
            if proportion > 0:
                sample_size = max(1, round(len(chars99) * proportion))
                chars_with_freq = [{'char': c, 'freq': freq_map.get(c, 0.0)} for c in chars99]
                chars_with_freq.sort(key=lambda x: x['freq'])
                sampled = []
                valid_sample = min(sample_size, len(chars_with_freq))
                for j in range(valid_sample):
                    start = int(j * len(chars_with_freq) / valid_sample)
                    end = int((j + 1) * len(chars_with_freq) / valid_sample)
                    stratum = chars_with_freq[start:end]
                    picked = stratum[len(stratum) // 2]
                    sampled.append(picked)
                groups.append({
                    'semester': UNCATEGORIZED,
                    'semesterName': '未收录',
                    'sampleSize': len(sampled),
                    'totalSource': len(chars99),
                    'proportion': proportion,
                    'chars': sampled,
                })
                total += len(sampled)

    return {'totalChars': total, 'groups': groups}


# ------------------------------------------------------------------
# 识读率推断
# ------------------------------------------------------------------
def build_stratum_lookup(test_result, answers):
    """
    根据测验结果和孩子的回答，建立层的认识状态查找表。
    返回 {(semester, stratum_index): bool}，bool 表示该层是否被认识。
    """
    lookup = {}
    for g in test_result['groups']:
        semester = g['semester']
        chars = g['chars']  # 已按频率排序且分层抽样
        # 这里需要重新计算每个抽样字属于哪一层
        # 简化：每个抽样字就是一层，按顺序对应 stratum_index 0,1,2,...
        for idx, item in enumerate(chars):
            char = item['char']
            known = answers.get(char, False)
            lookup[(semester, idx)] = known
    return lookup


def estimate_text_recognition_rate(text_content, test_result, answers, char_map, freq_map):
    """
    估算单篇文本的识读率。
    返回 0-1 之间的浮点数。
    """
    chars = extract_chinese_chars(text_content)
    if not chars:
        return 0.0

    # 建立层查找表
    lookup = build_stratum_lookup(test_result, answers)

    known_count = 0
    unknown_layers = set()

    for c in chars:
        sem = char_map.get(c, UNCATEGORIZED)
        freq = freq_map.get(c, 0.0)

        # 找到该字在对应学期的频率排序中的位置，从而确定层
        # 需要先获取该学期的所有字并按频率排序
        # 为了效率，这里我们重建一次该学期的分层信息
        if sem == UNCATEGORIZED:
            # 未收录字：查找是否有同层抽样结果
            known = False  # 保守估计
        else:
            # 找到该学期对应的组
            group = None
            for g in test_result['groups']:
                if g['semester'] == sem:
                    group = g
                    break
            if not group:
                known = False
            else:
                # 获取该学期所有源字并按频率排序
                # 但 test_result 中没有保存源字列表，这里需要重新构建
                # 为了简化，我们直接用该字的频率在组内做近似
                group_chars = group['chars']
                # 按频率排序后的抽样字列表
                sorted_sampled = sorted(group_chars, key=lambda x: x['freq'])
                # 找到该字应该属于哪一层
                # 用频率位置近似：把频率范围分成 sampleSize 层
                if len(sorted_sampled) == 0:
                    known = False
                else:
                    sample_size = len(sorted_sampled)
                    # 确定该字的层索引：按频率在 [min, max] 中的位置
                    freqs = [x['freq'] for x in sorted_sampled]
                    min_f, max_f = min(freqs), max(freqs)
                    if max_f == min_f:
                        stratum_idx = 0
                    else:
                        stratum_idx = int((freq - min_f) / (max_f - min_f) * (sample_size - 1))
                        stratum_idx = min(stratum_idx, sample_size - 1)
                    known = lookup.get((sem, stratum_idx), False)

        if known:
            known_count += 1

    return known_count / len(chars)


# ------------------------------------------------------------------
# 推荐逻辑
# ------------------------------------------------------------------
def recommend_text(texts, rates):
    """
    推荐单篇文本 + 整组反馈。
    texts: 10篇文本信息
    rates: 对应的识读率列表
    返回 {recommended, pass_count, total, rates_with_texts}
    """
    passed = [(t, r) for t, r in zip(texts, rates) if r >= PASS_THRESHOLD]
    passed.sort(key=lambda x: x[1], reverse=True)

    if passed:
        recommended = passed[0]  # 识读率最高且达标的文本
    else:
        # 没有达标，推荐识读率最高的一篇
        all_sorted = sorted(zip(texts, rates), key=lambda x: x[1], reverse=True)
        recommended = all_sorted[0]

    return {
        'recommended': recommended[0],
        'recommended_rate': recommended[1],
        'pass_count': len(passed),
        'total': len(texts),
        'all_passed': passed,
        'all_rates': list(zip(texts, rates)),
    }


# ------------------------------------------------------------------
# 安全与工具函数
# ------------------------------------------------------------------
# 危机关键词（从 skill_危机转介.md 加载）
SKILL_CRISIS = None  # 延迟加载

CRISIS_KEYWORDS = ["不想活了", "活着没意思", "想死", "自杀"]
STOP_KEYWORDS = ["不想读了", "不读了", "太难了", "不想读了"]
SELF_NEG = ["好笨", "我太笨了", "我很笨", "我读不好"]


def init_skill_crisis():
    """初始化加载 skill_危机转介.md"""
    global SKILL_CRISIS
    if SKILL_CRISIS is None:
        SKILL_CRISIS = load_skill_crisis()


def check_safety(text):
    """
    简化版 safety_check。
    优先使用 skill_危机转介.md 中的关键词，否则使用硬编码默认值。
    """
    init_skill_crisis()

    lower = text.lower()

    # 使用 skill 文件中的关键词（优先）
    if SKILL_CRISIS:
        crisis_kw = SKILL_CRISIS.get('crisis_keywords', CRISIS_KEYWORDS)
        yellow_kw = SKILL_CRISIS.get('yellow_keywords', STOP_KEYWORDS + SELF_NEG)
    else:
        crisis_kw = CRISIS_KEYWORDS
        yellow_kw = STOP_KEYWORDS + SELF_NEG

    if any(k in lower for k in crisis_kw):
        return "red"
    if any(k in lower for k in yellow_kw):
        return "yellow"
    return "green"


def get_crisis_response():
    """获取危机转介的标准话术（确定性，非 LLM 生成）"""
    return (
        "[系统] 检测到你似乎很难过。我们先不读了，没关系的。\n"
        "如果你有需要，一定要告诉爸爸妈妈、老师或学校心理老师。\n"
        "也可以拨打心理援助热线：400-161-9995。"
    )


def semester_to_grade_name(semester):
    """学期数转口语化年级名称。"""
    if semester == 1:
        return "一年级上学期"
    if semester == 2:
        return "一年级下学期"
    return f"{SEMESTER_NAMES[semester - 1]}"


def get_level_for_semester(semester):
    """获取学期对应的等级编码。"""
    return SEMESTER_TO_LEVEL.get(semester, '6')


def get_level_hint(semester):
    """当等级与学期不完全匹配时，返回给孩子的提示。"""
    if semester <= 2:
        return "这些字对你来说可能有一点点难，不过没关系，我们就像闯关一样试试看，不认识就跳过。"
    if semester >= 7:
        return "这些字对你来说可能比较简单，但我们可以快速试试，找到最适合你的书。"
    return ""


def get_available_themes_for_level(level):
    """获取某等级下可选的口语化主题列表。"""
    themes = set()
    for theme_name, levels in THEME_OPTIONS.items():
        if level in levels:
            themes.add(theme_name)
    return sorted(themes)


def get_folder_for_theme(level, theme_name):
    """根据等级和主题获取文件夹名。"""
    levels = THEME_OPTIONS.get(theme_name, {})
    return levels.get(level)


def select_texts_for_child(semester, theme_name=None, library=None):
    """
    为儿童选择10篇文本。
    优先按学期定等级，再按主题选；主题不匹配则在该等级下随机。
    返回 (folder_name, texts, level)。
    """
    if library is None:
        library = load_text_library()

    level = get_level_for_semester(semester)
    level_folders = [f for f in library.keys() if f.startswith(level)]

    if not level_folders:
        # 如果该等级没有文件夹，降级到最近的等级
        all_levels = sorted(set(SEMESTER_TO_LEVEL.values()), key=lambda x: int(x.replace('-', '')))
        idx = all_levels.index(level) if level in all_levels else 0
        if idx > 0:
            level = all_levels[idx - 1]
        level_folders = [f for f in library.keys() if f.startswith(level)]

    # 如果有主题，优先匹配
    if theme_name:
        target_folder = get_folder_for_theme(level, theme_name)
        if target_folder and target_folder in level_folders:
            return target_folder, library[target_folder][:10], level

    # 否则在该等级下随机选一个文件夹
    chosen_folder = random.choice(level_folders)
    return chosen_folder, library[chosen_folder][:10], level


# ------------------------------------------------------------------
# ReadingBuddyV2：识读率测验 agent
# ------------------------------------------------------------------
class ReadingBuddyV2:
    def __init__(self, load_history_on_start=True):
        # 确保危机检测已初始化
        init_skill_crisis()

        self.system_prompt = load_prompt()
        self.messages = [{"role": "system", "content": self.system_prompt}]
        self.state = "START"
        self.semester = 6  # 默认三下
        self.level = '6'
        self.theme_name = None
        self.folder_name = None
        self.texts = []
        self.test_result = None
        self.answers = {}  # {字: True/False}
        self.current_char_index = 0
        self.consecutive_errors = 0
        self.total_errors = 0

        # 加载数据
        self.char_map = load_char_table()
        self.freq_map = load_freq_table()
        self.prop_map = load_prop_table()
        self.library = load_text_library()

        if load_history_on_start:
            past = load_history()
            if past and past[0].get("role") == "system":
                past[0] = self.messages[0]
                self.messages = past
                print(f"[系统] 已从 {history_file} 恢复历史对话。\n")

    def reply(self, user_input):
        """处理用户输入，返回 AI 回复。"""
        safety = check_safety(user_input)

        # 红灯：危机信号
        if safety == "red":
            msg = get_crisis_response()
            return msg, True

        # 黄灯：情绪挫败
        if safety == "yellow":
            self.messages.append({"role": "user", "content": user_input})
            reply = "听起来刚才有点难，你已经很努力啦。我们先休息一下，不着急。你想再试一次，还是今天先到这里？"
            self.messages.append({"role": "assistant", "content": reply})
            save_history(self.messages)
            return reply, True

        # 正常流程
        if self.state == "START":
            return self._handle_start(user_input)
        elif self.state == "ASK_SEMESTER":
            return self._handle_semester(user_input)
        elif self.state == "ASK_THEME":
            return self._handle_theme(user_input)
        elif self.state == "SHOW_CHARS":
            return self._handle_show_chars(user_input)
        elif self.state == "TESTING":
            return self._handle_testing(user_input)
        elif self.state == "END":
            return self._handle_end(user_input)
        else:
            self.messages.append({"role": "user", "content": user_input})
            reply = chat(self.messages)
            self.messages.append({"role": "assistant", "content": reply})
            save_history(self.messages)
            return reply, False

    def _handle_start(self, user_input):
        """开场：说明身份并询问年级。"""
        self.messages.append({"role": "user", "content": user_input})

        prompt_addition = (
            "【系统提示：你是阅读小伙伴。请用温柔简短的语气回应，并询问孩子是几年级。"
            "你默认面对三年级下学期的孩子，但仍需确认。不要提及工具或算法。】"
        )
        temp_messages = self.messages + [{"role": "system", "content": prompt_addition}]
        reply = chat(temp_messages)

        self.messages.append({"role": "assistant", "content": reply})
        self.state = "ASK_SEMESTER"
        save_history(self.messages)
        return reply, False

    def _handle_semester(self, user_input):
        """解析年级/学期，询问主题。"""
        # 尝试从输入中解析学期数
        semester = self._parse_semester(user_input)
        if semester is not None:
            self.semester = semester
            self.level = get_level_for_semester(semester)

        # 获取该等级下的可选主题
        themes = get_available_themes_for_level(self.level)
        theme_list = "、".join(themes)

        self.messages.append({"role": "user", "content": user_input})
        prompt_addition = (
            f"【系统提示：孩子现在是{semester_to_grade_name(self.semester)}。"
            f"请温柔地确认，然后询问他想读什么主题。可选主题有：{theme_list}。"
            f"如果孩子说的主题不在这几个里，就从这几个里帮他挑一个最接近的。"
            f"如果他说随便，就告诉他你帮他选一个。{get_level_hint(self.semester)}】"
        )
        temp_messages = self.messages + [{"role": "system", "content": prompt_addition}]
        reply = chat(temp_messages)

        self.messages.append({"role": "assistant", "content": reply})
        self.state = "ASK_THEME"
        save_history(self.messages)
        return reply, False

    def _handle_theme(self, user_input):
        """解析主题，选择10篇文本，生成测验字表。"""
        # 解析主题
        themes = get_available_themes_for_level(self.level)
        theme_name = self._parse_theme(user_input, themes)

        # 选择10篇文本
        self.folder_name, self.texts, self.level = select_texts_for_child(
            self.semester, theme_name, self.library
        )
        self.theme_name = FOLDER_TO_THEME.get(self.folder_name, theme_name)

        # 生成测验字表
        self.test_result = assemble_test(
            self.semester, self.texts, self.char_map, self.freq_map, self.prop_map
        )

        # 构建展示给孩子的字表
        char_list = self.test_result['all_chars']
        char_display = "、".join(char_list)

        self.messages.append({"role": "user", "content": user_input})
        prompt_addition = (
            f"【系统提示：已为孩子选择了《{self.folder_name}》这个主题里的10篇小故事，"
            f"并从中生成了一个认字测验，共{len(char_list)}个字。"
            f"请用温柔简短的语气告诉孩子：你准备了一些他感兴趣主题故事里出现的字，让他来认。"
            f"他会看着这些字，一个个读出来；遇到不认识的字就跳过，不用紧张。"
            f"读完了之后，让他把认识的所有字一起发给你。"
            f"不要展示字表，字表会由系统在你说完之后自动追加。】"
        )
        temp_messages = self.messages + [{"role": "system", "content": prompt_addition}]
        intro = chat(temp_messages)

        # 在代码里确保字表完整展示，避免 LLM 漏字或分批
        char_display = "、".join(char_list)
        reply = (
            f"{intro}\n\n"
            f"来，请你认一认这些字，认识的就读出来，不认识的就跳过：\n\n"
            f"{char_display}\n\n"
            f"读完了之后，把你认识的所有字一起发给我就可以啦，连在一起发也没关系哦！"
        )

        self.messages.append({"role": "assistant", "content": reply})
        self.state = "TESTING"
        save_history(self.messages)
        return reply, False

    def _handle_testing(self, user_input):
        """处理孩子的认读结果，推断识读率，推荐文本。"""
        # 解析孩子读出的字
        recognized_chars = self._parse_recognized_chars(user_input)

        # 记录每个测验字的回答
        for char in self.test_result['all_chars']:
            self.answers[char] = char in recognized_chars

        # 计算每篇文本的识读率
        rates = []
        for text in self.texts:
            rate = estimate_text_recognition_rate(
                text['content'], self.test_result, self.answers,
                self.char_map, self.freq_map
            )
            rates.append(rate)

        # 推荐
        rec = recommend_text(self.texts, rates)

        # 生成反馈
        pass_count = rec['pass_count']
        rec_text = rec['recommended']
        rec_rate = rec['recommended_rate']
        rec_title = extract_readable_title(rec_text['name'])

        self.messages.append({"role": "user", "content": user_input})
        prompt_addition = (
            f"【系统提示：孩子已经把他认识的所有字都发给你了（不认识的字他没有发）。"
            f"请直接基于这些结果给出反馈和推荐，不要再要求孩子继续读剩下的字，也不要反问孩子问题。"
            f"测验结果：共{len(self.answers)}个字，孩子认识了{sum(self.answers.values())}个。"
            f"10篇文本中，有{pass_count}篇的识读率达到95%以上。"
            f"推荐文本是《{rec_title}》（识读率约{rec_rate*100:.1f}%）。"
            f"无论识读率是否达到95%，你都请用孩子能听懂的话：肯定他的努力；"
            f"告诉他这组故事里他大概能读几本（或直接说今天推荐这一本）；"
            f"明确推荐《{rec_title}》作为今天可以试试读的故事。不要提及具体算法。】"
        )
        temp_messages = self.messages + [{"role": "system", "content": prompt_addition}]
        reply = chat(temp_messages)

        self.messages.append({"role": "assistant", "content": reply})
        self.state = "END"
        save_history(self.messages)
        return reply, False

    def _handle_end(self, user_input):
        """对话结束后的兜底回复。"""
        self.messages.append({"role": "user", "content": user_input})
        reply = chat(self.messages)
        self.messages.append({"role": "assistant", "content": reply})
        save_history(self.messages)
        return reply, False

    def _parse_semester(self, text):
        """从用户输入中解析学期数。"""
        # 匹配"三年级下学期"、"三下"、"3年级下学期"等
        patterns = [
            r'三上|三年级上',
            r'三下|三年级下',
            r'二上|二年级上',
            r'二下|二年级下',
            r'四上|四年级上',
            r'四下|四年级下',
            r'五上|五年级上',
            r'五下|五年级下',
            r'六上|六年级上',
            r'六下|六年级下',
            r'一上|一年级上',
            r'一下|一年级下',
        ]
        mapping = {
            '一上': 1, '一下': 2,
            '二上': 3, '二下': 4,
            '三上': 5, '三下': 6,
            '四上': 7, '四下': 8,
            '五上': 9, '五下': 10,
            '六上': 11, '六下': 12,
            '一年级上': 1, '一年级下': 2,
            '二年级上': 3, '二年级下': 4,
            '三年级上': 5, '三年级下': 6,
            '四年级上': 7, '四年级下': 8,
            '五年级上': 9, '五年级下': 10,
            '六年级上': 11, '六年级下': 12,
        }
        for pat in patterns:
            m = re.search(pat, text)
            if m:
                return mapping.get(m.group(), None)

        # 尝试直接匹配数字"6"、"三年级"等
        if re.search(r'三年级', text):
            return 6
        if re.search(r'二年级', text):
            return 4
        if re.search(r'一年级', text):
            return 2
        if re.search(r'四年级', text):
            return 8
        if re.search(r'五年级', text):
            return 10
        if re.search(r'六年级', text):
            return 12

        # 尝试匹配"学期6"、"6学期"等
        m = re.search(r'学期\s*(\d+)|(\d+)\s*学期', text)
        if m:
            sem = int(m.group(1) or m.group(2))
            if 1 <= sem <= 12:
                return sem
        return None

    def _parse_theme(self, text, available_themes):
        """从用户输入中解析主题。"""
        # 直接匹配
        for theme in available_themes:
            if theme in text:
                return theme

        # 关键词映射
        keyword_map = {
            "童话故事和诗歌": ["童话", "诗歌", "诗", "儿歌", "故事"],
            "生活中的小故事和神话寓言": ["神话", "寓言", "生活故事", "小故事", "故事"],
            "生活故事和散文": ["散文", "生活", "故事"],
            "生活故事和历史故事": ["历史", "历史故事", "故事"],
        }
        for theme, keywords in keyword_map.items():
            if theme in available_themes:
                for kw in keywords:
                    if kw in text:
                        return theme

        # 如果没匹配到，随机返回一个可用主题
        return random.choice(available_themes)

    def _parse_recognized_chars(self, text):
        """从用户输入中解析出孩子读出的汉字。"""
        return set(extract_chinese_chars(text))


# ------------------------------------------------------------------
# 启动
# ------------------------------------------------------------------
if __name__ == "__main__":
    print("[系统] 小航的阅读小伙伴 Agent V2 已启动（识读率测验版）。")
    print("[系统] 输入 /clear 清空历史，exit / quit / 退出 结束对话。\n")

    buddy = ReadingBuddyV2()

    while True:
        try:
            user_input = input("你：").strip()
        except EOFError:
            print("\n[系统] 检测到输入结束，保存并退出。")
            save_history(buddy.messages)
            break

        if not user_input:
            continue

        if user_input.lower() in {"exit", "quit", "退出"}:
            print("\n[系统] 再见！")
            save_history(buddy.messages)
            break

        if user_input == "/clear":
            buddy.messages = [{"role": "system", "content": buddy.system_prompt}]
            buddy.state = "START"
            print("[系统] 历史已清空。\n")
            continue

        reply, is_system = buddy.reply(user_input)

        if is_system:
            print(reply + "\n")
        else:
            print(f"AI：{reply}\n")

# MIT License | 郑先隽，北师大心理学部教授，人本 AI 设计与创新

